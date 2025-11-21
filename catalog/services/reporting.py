# catalog/services/reporting.py
from dataclasses import dataclass
from typing import Iterable, List, Protocol
import io, csv

try:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.pdfgen import canvas
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False


@dataclass
class ReportColumn:
    header: str
    accessor: str  # clave en cada fila de datos


class ReportRenderer(Protocol):
    def render(self, title: str, columns: List[ReportColumn], rows: Iterable[dict]) -> bytes: ...


class CSVRenderer:
    content_type = "text/csv"
    extension = "csv"

    def render(self, title, columns, rows) -> bytes:
        buf = io.StringIO(newline="")
        writer = csv.writer(buf)
        writer.writerow([c.header for c in columns])
        for r in rows:
            writer.writerow([r.get(c.accessor, "") for c in columns])
        return buf.getvalue().encode("utf-8")


class PDFRenderer:
    content_type = "application/pdf"
    extension = "pdf"

    def render(self, title, columns, rows) -> bytes:
        if not REPORTLAB_OK:
            raise RuntimeError("ReportLab no está instalado para PDF.")
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=LETTER)
        width, height = LETTER

        y = height - 36
        c.setFont("Helvetica-Bold", 14)
        c.drawString(36, y, title)
        y -= 24

        c.setFont("Helvetica-Bold", 10)
        x = 36
        for col in columns:
            c.drawString(x, y, col.header)
            x += 150
        y -= 18

        c.setFont("Helvetica", 10)
        for r in rows:
            x = 36
            if y < 36:
                c.showPage()
                y = height - 36
            for col in columns:
                txt = str(r.get(col.accessor, ""))
                c.drawString(x, y, txt[:25])
                x += 150
            y -= 16

        c.showPage()
        c.save()
        return buf.getvalue()


class ExcelRenderer:
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    def render(self, title, columns, rows) -> bytes:
        if not OPENPYXL_OK:
            raise RuntimeError("openpyxl no está instalado para Excel.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"

        ws.append([c.header for c in columns])
        for r in rows:
            ws.append([r.get(c.accessor, "") for c in columns])

        # ancho de columnas
        for i, col in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(col.header) + 2)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


class DefaultReportFactory:
    @staticmethod
    def for_format(fmt: str) -> ReportRenderer:
        fmt = (fmt or "").lower()
        if fmt == "pdf":
            return PDFRenderer()
        if fmt in ("xlsx", "excel"):
            return ExcelRenderer()
        # fallback
        return CSVRenderer()
